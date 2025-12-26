from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple, Optional, Any

from openpyxl import Workbook, load_workbook

import xlrd
from xlrd.xldate import xldate_as_datetime


colums_head = ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]


class OpenPyXLWorkbook:
    def __init__(self, path: Path, *, read_only: bool, data_only: bool):
        self.path = path
        self.read_only = read_only
        self.data_only = data_only
        self._wb: Optional[object] = None

    def __enter__(self):
        self._wb = load_workbook(
            filename=str(self.path),
            read_only=self.read_only,
            data_only=self.data_only,
        )
        return self._wb

    def __exit__(self, exc_type, exc, tb):
        if self._wb is not None:
            self._wb.close()
        return False


class XLRDBook:
    def __init__(self, path: Path):
        self.path = path
        self._book: Optional[xlrd.book.Book] = None

    def __enter__(self) -> xlrd.book.Book:
        self._book = xlrd.open_workbook(str(self.path))
        return self._book

    def __exit__(self, exc_type, exc, tb):
        if self._book is not None:
            self._book.release_resources()
        return False


class NewWorkbook:
    def __init__(self):
        self._wb: Optional[Workbook] = None

    def __enter__(self) -> Workbook:
        self._wb = Workbook()
        return self._wb

    def __exit__(self, exc_type, exc, tb):
        if self._wb is not None:
            self._wb.close()
        return False


def norm_text(value: object) -> str:
    """Преобразует значение из ячейки Excel в строку без пробелов.

    Args:
        value: Значение из ячейки Excel

    Returns:
         Строковое значение (str)
    """
    return "" if value is None else str(value).strip()


def casefold_text(value: object) -> str:
    """Возвращает строковый ключ (str) для сравнения без учёта регистра.

    Args:
        value: Значение из ячейки Excel любого типа.

    Returns:
        Строковый ключ (str) для сравнения:
    """
    return norm_text(value).casefold()


def get_excel_format(path: Path) -> str:
    """Определяет библиотеку для чтения Excel по расширению файла.

    Args:
        path: Путь к Excel-файлу.

    Returns:
        Строковый идентификатор:
        - "xlrd" для .xls
        - "openpyxl" для .xlsx и .xlsm
    """
    format_exel = path.suffix.lower()
    if format_exel == ".xls":
        return "xlrd"
    if format_exel in (".xlsx", ".xlsm"):
        return "openpyxl"
    raise ValueError(f"Неподдерживаемый формат файла: {format_exel}. "
                     f"Нужно .xls/.xlsx/.xlsm")


def map_columns_to_indexes(headers: List[str]) -> Dict[str, int]:
    """Записывает'название колонки -> индекс'.

    Args:
        headers: Список заголовков из строки шапки таблицы.

    Returns:
        Словарь, где ключ — название колонки, значение — индекс
    """
    return {name: idx for idx, name in enumerate(headers) if name}


def find_header_row_from_rows(rows: Iterable[List[object]],
                              scan_rows: int = 50) -> Tuple[int, List[str]]:
    """
    Ищет строку заголовков таблицы в первых 50 строках файла.

    Args:
        rows: Итератор строк (список значений ячеек строки).
        scan_rows: Сколько первых строк анализировать.(50)
    Returns:
        Tuple (номер строки, список_заголовков_строки).
    """
    required_norm = {casefold_text(column_name) for column_name in colums_head}

    best_row_number = 1
    best_headers: List[str] = []
    best_score = -1

    for row_number, row_values in enumerate(rows, start=1):
        if row_number > scan_rows:
            break

        headers = [norm_text(cell_value) for cell_value in row_values]
        header_norm = {casefold_text(header) for header in headers if header}

        score = len(required_norm.intersection(header_norm))
        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_headers = headers

        if score == len(required_norm):
            return row_number, headers

    return best_row_number, best_headers


def read_xls_cell_value(book: xlrd.book.Book,
                        sheet: xlrd.sheet.Sheet,
                        row_index: int,
                        col_index: int) -> Any:
    """Читает значение ячейки из .xls и приводит типы к читаемому виду из за старого формата.

        Args:
            book: Открытая книга xlrd
            sheet: Лист excel xlrd
            row_index: Индекс строки .
            col_index: Индекс столбца.

        Returns:
            Значение ячейки (object). Возможные типы: datetime, int, float, str, None и т.д.
        """
    cell = sheet.cell(row_index, col_index)

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xldate_as_datetime(cell.value, book.datemode)

    if cell.ctype == xlrd.XL_CELL_NUMBER:
        if float(cell.value).is_integer():
            return int(cell.value)
        return cell.value

    return cell.value


@dataclass(frozen=True)
class FilterRequest:
    input_path: Path
    filter_column: str
    filter_value: str
    output_path: Optional[Path] = None  # нужен только на шаге сохранения (если захочешь)


class ExcelService:
    def read_headers(self, exel_path: Path) -> List[str]:
        """Читает заголовки таблицы из Excel-файла.

        Метод ищет строку заголовков в первых строках файла.

        Args:
            exel_path: Путь к Excel-файлу.

        Returns:
            Список заголовков (названий столбцов).
        """
        backend = get_excel_format(exel_path)

        if backend == "openpyxl":
            with OpenPyXLWorkbook(exel_path, read_only=True, data_only=True) as workbook:
                worksheet = workbook.active
                first_rows = ([cell.value for cell in row] for row in worksheet.iter_rows(min_row=1, max_row=50))
                header_row_number, headers = find_header_row_from_rows(first_rows)
                return [header for header in headers if header]

        with XLRDBook(exel_path) as book:
            sheet = book.sheet_by_index(0)
            first_rows = (sheet.row_values(row_index) for row_index in range(min(50, sheet.nrows)))
            header_row_number, headers = find_header_row_from_rows(first_rows)
            return [header for header in headers if header]

    def get_unique_values(self, exel_path: Path, column_name: str, limit: int = 500) -> List[str]:
        """Возвращает уникальные значения из выбранного столбца (без учёта регистра).


           Args:
               file_path: Путь к Excel-файлу.
               column_name: Название столбца, из которого нужно получить значения.
               limit: Максимальное количество уникальных значений, которое вернём.

           Returns:
               Список уникальных строковых значений столбца

           """
        backend = get_excel_format(exel_path)

        if backend == "openpyxl":
            with OpenPyXLWorkbook(exel_path, read_only=True, data_only=True) as  workbook:
                worksheet =  workbook.active
                rows_for_detect = ([cell.value for cell in row] for row in worksheet.iter_rows(min_row=1, max_row=50))
                header_row_num, raw_headers = find_header_row_from_rows(rows_for_detect)

                headers = [norm_text(header_cell) for header_cell in raw_headers]
                column_to_index = map_columns_to_indexes(headers)

                if column_name not in column_to_index:
                    raise ValueError(f"Колонка не найдена: '{column_name}'")

                target_col_index = column_to_index[column_name]
                seen: set[str] = set()
                result_values: List[str] = []

                data_rows = worksheet.iter_rows(
                    min_row=header_row_num + 1,
                    values_only=True,
                )

                for row_values in data_rows:
                    if target_col_index >= len(row_values):
                        continue

                    raw_value = row_values[target_col_index]
                    value_text = norm_text(raw_value)
                    if not value_text:
                        continue

                    value_key = value_text.casefold()
                    if value_key in seen:
                        continue

                    seen.add(value_key)
                    result_values.append(value_text)

                    if len(result_values) >= limit:
                            break

                return result_values

        with XLRDBook(exel_path) as book:
            sheet = book.sheet_by_index(0)

            first_rows = (sheet.row_values(row_index) for row_index in range(min(50, sheet.nrows)))
            header_row_number, raw_headers = find_header_row_from_rows(first_rows)

            headers = [norm_text(header_cell) for header_cell in raw_headers]
            column_to_index = map_columns_to_indexes(headers)

            if column_name not in column_to_index:
                raise ValueError(f"Колонка не найдена: '{column_name}'")

            target_col_index = column_to_index[column_name]

            seen_keys: set[str] = set()
            result_values: List[str] = []

            start_row_index = header_row_number
            for row_index in range(start_row_index, sheet.nrows):
                if target_col_index >= sheet.ncols:
                    continue

                raw_value = read_xls_cell_value(book, sheet, row_index, target_col_index)
                value_text = norm_text(raw_value)
                if not value_text:
                    continue

                value_key = value_text.casefold()
                if value_key in seen_keys:
                    continue

                seen_keys.add(value_key)
                result_values.append(value_text)

                if len(result_values) >= limit:
                    break

            return result_values

    def filter_rows(self, request: FilterRequest) -> Tuple[List[str], List[List[Any]]]:
        """Фильтрует строки входного Excel и возвращает данные для сохранения.

        Args:
            request: Параметры фильтрации.

        Returns:
                headers: Заголовки выходного файла.
                rows: Отфильтрованные строки.
        """
        if not request.input_path.exists():
            raise FileNotFoundError(f"Входной файл не найден: {request.input_path}")

        engine = get_excel_format(request.input_path)
        target_value_key = casefold_text(request.filter_value)

        if engine == "openpyxl":
            with OpenPyXLWorkbook(request.input_path, read_only=True, data_only=True) as input_workbook:
                worksheet = input_workbook.active

                first_rows = (
                    [cell.value for cell in row]
                    for row in worksheet.iter_rows(min_row=1, max_row=50)
                )
                header_row_number, raw_headers = find_header_row_from_rows(first_rows)

                headers = [norm_text(header_cell) for header_cell in raw_headers]
                column_to_index = map_columns_to_indexes(headers)

                self.validate_columns(column_to_index, request.filter_column)

                filter_col_index = column_to_index[request.filter_column]
                keep_col_indices = [column_to_index[col] for col in colums_head]

                result_rows: List[List[Any]] = []
                for row_values in worksheet.iter_rows(min_row=header_row_number + 1, values_only=True):
                    cell_value = row_values[filter_col_index] if filter_col_index < len(row_values) else None
                    if casefold_text(cell_value) != target_value_key:
                        continue

                    result_rows.append(
                        [row_values[index] if index < len(row_values) else None for index in keep_col_indices]
                    )

                return colums_head, result_rows

        with XLRDBook(request.input_path) as book:
            sheet = book.sheet_by_index(0)

            first_rows = (sheet.row_values(row_index) for row_index in range(min(50, sheet.nrows)))
            header_row_number, raw_headers = find_header_row_from_rows(first_rows)

            headers = [norm_text(header_cell) for header_cell in raw_headers]
            column_to_index = map_columns_to_indexes(headers)

            self.validate_columns(column_to_index, request.filter_column)

            filter_col_index = column_to_index[request.filter_column]
            keep_col_indices = [column_to_index[col] for col in colums_head]

            result_rows: List[List[Any]] = []
            start_row_index = header_row_number

            for row_index in range(start_row_index, sheet.nrows):
                cell_value = (
                    read_xls_cell_value(book, sheet, row_index, filter_col_index)
                    if filter_col_index < sheet.ncols
                    else None
                )
                if casefold_text(cell_value) != target_value_key:
                    continue

                output_row: List[Any] = []
                for keep_index in keep_col_indices:
                    output_row.append(
                        read_xls_cell_value(book, sheet, row_index, keep_index)
                        if keep_index < sheet.ncols
                        else None
                    )

                result_rows.append(output_row)

            return colums_head, result_rows

    def save_xlsx(self, output_path: Path, headers: List[str], rows: List[List[Any]]) -> None:
        """Сохраняет результат в новый файл.

        Args:
            output_path: Путь сохранения файла.
            headers: Заголовки
            rows: Данные
        """
        if output_path.suffix.lower() != ".xlsx":
            raise ValueError("Файл результата должен быть .xlsx")

        with NewWorkbook() as output_workbook:
            output_sheet = output_workbook.active
            output_sheet.title = "Отфильтрованно"

            output_sheet.append(headers)
            for row_values in rows:
                output_sheet.append(row_values)

            output_workbook.save(str(output_path))

    @staticmethod
    def validate_columns(column_to_index: Dict[str, int], filter_column: str) -> None:
        """Проверяет наличие обязательных колонок и колонки фильтра.

        Args:
            column_to_index: Словарь "название колонки".
            filter_column: Название колонки, по которой будет выполняться фильтрация.

        """
        if filter_column not in column_to_index:
            raise ValueError(f"Колонка для фильтра не найдена: '{filter_column}'")

        missing_columns = [col for col in colums_head if col not in column_to_index]
        if missing_columns:
            raise ValueError("В исходном файле отсутствуют колонки: " + ", ".join(missing_columns))